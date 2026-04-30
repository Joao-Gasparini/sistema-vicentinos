from dotenv import load_dotenv

load_dotenv()

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from datetime import datetime
from PIL import Image, UnidentifiedImageError
from email_validator import validate_email, EmailNotValidError
from phonenumbers import NumberParseException
from flask_wtf.csrf import CSRFProtect
from functools import wraps

import phonenumbers
import time
import re
import os
import io

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from config import Config
from models import db, Vicentino, Conselho, Conferencia, Familia, Atendimento

# =========================================================
# INICIALIZAÇÃO DA APLICAÇÃO
# =========================================================

app = Flask(__name__)
app.config.from_object(Config)

# Inicializa a proteção CSRF em todos os formulários
csrf = CSRFProtect(app)

db.init_app(app)
mail = Mail(app)
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# Define e cria o diretório para armazenar as fotos de perfil dos usuários
app.config['UPLOAD_FOLDER_USUARIO'] = os.path.join(app.root_path, 'static', 'fotos_perfil')
os.makedirs(app.config['UPLOAD_FOLDER_USUARIO'], exist_ok=True)

# Limite máximo de upload: 5MB
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024


@app.errorhandler(413)
def arquivo_muito_grande(e):
    """Captura o erro 413 (arquivo maior que o limite) e exibe mensagem amigável ao usuário."""
    flash('Arquivo muito grande! Máximo permitido: 5MB.', 'danger')
    return redirect(request.referrer or url_for('home'))


# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================

def apenas_letras(texto):
    """Verifica se o texto contém apenas letras (incluindo acentuadas) e espaços."""
    return re.fullmatch(r"[A-Za-zÀ-ÿ\s]+", texto) is not None


def email_valido(email):
    """
    Valida e normaliza um endereço de e-mail.
    Retorna o e-mail normalizado se válido, ou None se inválido.
    """
    try:
        valid = validate_email(email, check_deliverability=False)
        return valid.normalized
    except EmailNotValidError:
        return None


def telefone_valido_br(telefone):
    """
    Valida um número de telefone brasileiro e o retorna no formato E.164 (+55...).
    Retorna None se o número for inválido.
    """
    try:
        numero = phonenumbers.parse(telefone, 'BR')

        if phonenumbers.is_valid_number(numero):
            return phonenumbers.format_number(
                numero,
                phonenumbers.PhoneNumberFormat.E164
            )

        return None

    except NumberParseException:
        return None


def telefone_para_input(telefone_e164):
    """
    Converte um telefone no formato E.164 (+55...) para apenas os dígitos locais,
    removendo o código do país (+55). Usado para preencher campos de formulário.
    """
    if not telefone_e164:
        return ''

    # Remove o prefixo +55
    telefone = telefone_e164.replace('+55', '')

    return telefone  # Retorna só os números (ex: 11999998888)

# =========================================================
# DOCUMENTOS (CPF / CNPJ)
# =========================================================

def limpar_numero(valor):
    """Remove qualquer caractere que não seja número."""
    return re.sub(r'\D', '', valor)


def cpf_valido(cpf):
    """
    Valida um CPF brasileiro verificando:
    - Se possui 11 dígitos
    - Se não é sequência repetida
    - Se os dígitos verificadores são válidos
    """
    cpf = limpar_numero(cpf)

    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    # 1º dígito
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    resto = (soma * 10) % 11
    resto = 0 if resto == 10 else resto
    if resto != int(cpf[9]):
        return False

    # 2º dígito
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    resto = (soma * 10) % 11
    resto = 0 if resto == 10 else resto
    if resto != int(cpf[10]):
        return False

    return True


def cnpj_valido(cnpj):
    """
    Valida um CNPJ brasileiro verificando:
    - Se possui 14 dígitos
    - Se não é sequência repetida
    - Se os dígitos verificadores são válidos
    """
    cnpj = limpar_numero(cnpj)

    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    pesos2 = [6] + pesos1

    soma = sum(int(cnpj[i]) * pesos1[i] for i in range(12))
    dig1 = 11 - (soma % 11)
    dig1 = dig1 if dig1 < 10 else 0

    soma = sum(int(cnpj[i]) * pesos2[i] for i in range(13))
    dig2 = 11 - (soma % 11)
    dig2 = dig2 if dig2 < 10 else 0

    return cnpj[-2:] == f"{dig1}{dig2}"


def documento_valido(documento, tipo_usuario):
    """
    Valida CPF ou CNPJ dependendo do tipo do usuário.
    - admin → CNPJ
    - vicentino → CPF
    """
    documento = limpar_numero(documento)

    if not documento:
        return False

    if tipo_usuario == 'admin':
        return cnpj_valido(documento)
    else:
        return cpf_valido(documento)

def gerar_token(email, salt):
    """Gera um token seguro e com tempo de expiração baseado no e-mail e em um salt específico."""
    return serializer.dumps(email, salt=salt)


def admin_required(f):
    """
    Decorator que protege rotas de acesso exclusivo a administradores.
    Redireciona para o login se o usuário não estiver autenticado,
    ou para a home se não for admin.
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'vicentino_id' not in session:
            return redirect(url_for('login'))

        usuario = Vicentino.query.get(session['vicentino_id'])

        if usuario.tipo != 'admin':
            flash('Acesso restrito.', 'danger')
            return redirect(url_for('home'))

        return f(*args, **kwargs)

    return decorated_function


def validar_token(token, salt, expiracao):
    """
    Valida um token seguro e retorna o e-mail contido nele.
    Retorna None se o token estiver expirado ou for inválido.
    """
    try:
        email = serializer.loads(token, salt=salt, max_age=expiracao)
        return email
    except SignatureExpired:
        return None
    except BadSignature:
        return None


def enviar_email_confirmacao(vicentino):
    """
    Gera um token de confirmação e envia um e-mail ao vicentino
    com o link para ativar sua conta.
    """
    token = gerar_token(vicentino.email, 'confirmar-email')
    link = url_for('confirmar_email', token=token, _external=True)

    msg = Message(
        subject='Confirmação de cadastro - Famílias Assistidas SSVP',
        recipients=[vicentino.email]
    )

    msg.body = f'''Olá, {vicentino.nome}!

Seu cadastro foi realizado com sucesso.

Para confirmar seu e-mail e ativar sua conta, clique no link abaixo:

{link}

Se você não solicitou este cadastro, ignore este e-mail.
'''

    mail.send(msg)


def enviar_email_redefinicao(vicentino):
    """
    Gera um token de redefinição e envia um e-mail ao vicentino
    com o link para cadastrar uma nova senha.
    """
    token = gerar_token(vicentino.email, 'redefinir-senha')
    link = url_for('redefinir_senha', token=token, _external=True)

    msg = Message(
        subject='Redefinição de senha - Famílias Assistidas SSVP',
        recipients=[vicentino.email]
    )

    msg.body = f'''Olá, {vicentino.nome}!

Recebemos uma solicitação para redefinir sua senha.

Para cadastrar uma nova senha, clique no link abaixo:

{link}

Se você não solicitou essa alteração, ignore este e-mail.
'''

    mail.send(msg)


# =========================================================
# RATE LIMITING DE LOGIN (in-memory, por IP)
# =========================================================

_login_tentativas = {}       # { ip: [timestamps de tentativas falhas] }
_LOGIN_MAX_TENTATIVAS = 5    # Máximo de falhas permitidas
_LOGIN_JANELA_SEGUNDOS = 300 # Janela de 5 minutos


def _login_bloqueado(ip):
    """Retorna True se o IP excedeu o limite de tentativas."""
    agora = time.time()
    tentativas = [t for t in _login_tentativas.get(ip, []) if agora - t < _LOGIN_JANELA_SEGUNDOS]
    _login_tentativas[ip] = tentativas
    return len(tentativas) >= _LOGIN_MAX_TENTATIVAS


def _registrar_falha_login(ip):
    """Registra uma tentativa de login mal-sucedida para o IP."""
    agora = time.time()
    tentativas = [t for t in _login_tentativas.get(ip, []) if agora - t < _LOGIN_JANELA_SEGUNDOS]
    tentativas.append(agora)
    _login_tentativas[ip] = tentativas


# =========================================================
# ROTAS PÚBLICAS / INICIAIS
# =========================================================

@app.route('/')
def home():
    """Redireciona para o dashboard se logado, ou para o login se não autenticado."""
    if 'vicentino_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    cpf_input = ''

    if request.method == 'POST':
        ip = request.remote_addr
        if _login_bloqueado(ip):
            flash('Muitas tentativas de login. Aguarde 5 minutos e tente novamente.', 'danger')
            return render_template('login.html', cpf_input='')

        cpf_input = request.form.get('cpf', '').strip()
        senha = request.form.get('senha', '')

        documento = re.sub(r'\D', '', cpf_input)

        if not documento or not senha:
            flash('Informe o documento e a senha.', 'danger')
            return render_template('login.html', cpf_input=cpf_input)

        # Decide se é CPF (11 dígitos) ou CNPJ (14 dígitos)
        if len(documento) == 11:
            if not cpf_valido(documento):
                flash('CPF inválido.', 'danger')
                return render_template('login.html', cpf_input=cpf_input)
            vicentino = Vicentino.query.filter_by(cpf=documento).first()

        elif len(documento) == 14:
            vicentino = Vicentino.query.filter_by(cnpj=documento).first()

        else:
            flash('Documento inválido. Informe um CPF ou CNPJ válido.', 'danger')
            return render_template('login.html', cpf_input=cpf_input)

        if not vicentino or not check_password_hash(vicentino.senha_hash, senha):
            _registrar_falha_login(ip)
            flash('Documento ou senha incorretos.', 'danger')
            return render_template('login.html', cpf_input=cpf_input)

        session['vicentino_id'] = vicentino.id
        session['vicentino_nome'] = vicentino.nome
        session['vicentino_email'] = vicentino.email
        session['vicentino_foto'] = vicentino.foto if hasattr(vicentino, 'foto') else None
        session['tipo'] = vicentino.tipo

        flash('Login realizado com sucesso.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('login.html', cpf_input='')


@app.route('/api/conferencias/<int:conselho_id>')
@admin_required
def api_conferencias(conselho_id):
    conferencias = Conferencia.query.filter_by(conselho_id=conselho_id).order_by(Conferencia.nome).distinct().all()
    return jsonify([{'id': c.id, 'nome': c.nome} for c in conferencias])

@app.route('/admin/cadastrar_vicentino', methods=['GET', 'POST'])
@admin_required
def cadastrar_vicentino():
    conselhos = Conselho.query.all()
    conferencias = Conferencia.query.all()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        cpf_input = request.form.get('cpf', '').strip()
        cpf = re.sub(r'\D', '', cpf_input)
        email_informado = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        senha = request.form.get('senha', '')
        confirmar_senha = request.form.get('confirmar_senha', '')
        foto = request.files.get('foto')

        # ← agora vêm como IDs numéricos
        conselho_id = request.form.get('conselho_id') or None
        conferencia_id = request.form.get('conferencia_id') or None

        if email_informado == '':
            email = None
        else:
            email = email_valido(email_informado)
            if not email:
                flash('Informe um e-mail válido.', 'danger')
                return redirect(url_for('cadastrar_vicentino'))

        if not nome or not sobrenome or not senha or not confirmar_senha:
            flash('Preencha todos os campos obrigatórios.', 'danger')
            return redirect(url_for('cadastrar_vicentino'))

        if not apenas_letras(nome) or not apenas_letras(sobrenome):
            flash('Nome e sobrenome devem conter apenas letras.', 'danger')
            return redirect(url_for('cadastrar_vicentino'))

        if cpf:
            if not cpf_valido(cpf):
                flash('CPF inválido.', 'danger')
                return redirect(url_for('cadastrar_vicentino'))
            if Vicentino.query.filter_by(cpf=cpf).first():
                flash('Já existe um usuário com esse CPF.', 'danger')
                return redirect(url_for('cadastrar_vicentino'))

        if len(senha) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'danger')
            return redirect(url_for('cadastrar_vicentino'))

        if senha != confirmar_senha:
            flash('As senhas não coincidem.', 'danger')
            return redirect(url_for('cadastrar_vicentino'))

        if email:
            if Vicentino.query.filter_by(email=email).first():
                flash('Já existe um vicentino cadastrado com este e-mail.', 'danger')
                return redirect(url_for('cadastrar_vicentino'))

        if email:
            status = 'pendente'
            email_confirmado = False
        else:
            status = 'ativo'
            email_confirmado = True

        novo_vicentino = Vicentino(
            nome=nome,
            sobrenome=sobrenome,
            cpf=cpf if cpf else None,
            email=email,
            telefone=telefone,
            senha_hash=generate_password_hash(senha),
            status=status,
            email_confirmado=email_confirmado,
            tipo='vicentino',
            foto=None,
            conselho_id=conselho_id,       # ← FK
            conferencia_id=conferencia_id  # ← FK
        )

        db.session.add(novo_vicentino)
        db.session.commit()

        if email:
            try:
                enviar_email_confirmacao(novo_vicentino)
                flash('Vicentino cadastrado! E-mail de confirmação enviado.', 'success')
            except Exception:
                flash('Vicentino cadastrado, mas erro ao enviar e-mail.', 'warning')
        else:
            flash('Vicentino cadastrado com sucesso!', 'success')

        return redirect(url_for('dashboard'))

    return render_template(
        'cadastro_vicentino.html',
        conselhos=conselhos,
        conferencias=conferencias
    )


@app.route('/admin/familias')
@admin_required
def listar_familias():
    """Lista todas as famílias assistidas com filtros. Apenas admin."""
    nome_busca       = request.args.get('nome', '').strip()
    conferencia_raw  = request.args.get('conferencia_id', '')
    vicentino_raw    = request.args.get('vicentino_id', '')
    status_filtro    = request.args.get('status', '')

    query = Familia.query

    if nome_busca:
        query = query.filter(Familia.nome_responsavel.ilike(f'%{nome_busca}%'))

    if conferencia_raw:
        try:
            query = query.filter_by(conferencia_id=int(conferencia_raw))
        except (ValueError, TypeError):
            pass

    if vicentino_raw:
        try:
            query = query.filter_by(vicentino_id=int(vicentino_raw))
        except (ValueError, TypeError):
            pass

    todas = query.order_by(Familia.nome_responsavel).all()

    if status_filtro == 'ativa':
        ativas   = [f for f in todas if f.status == 'ativa']
        inativas = []
    elif status_filtro == 'inativa':
        ativas   = []
        inativas = [f for f in todas if f.status == 'inativa']
    else:
        ativas   = [f for f in todas if f.status == 'ativa']
        inativas = [f for f in todas if f.status == 'inativa']

    conferencias = Conferencia.query.order_by(Conferencia.nome).all()
    vicentinos   = Vicentino.query.filter_by(tipo='vicentino').order_by(Vicentino.nome).all()

    return render_template(
        'admin_familias.html',
        ativas=ativas,
        inativas=inativas,
        conferencias=conferencias,
        vicentinos=vicentinos,
    )


@app.route('/admin/familia/<int:familia_id>/toggle_status', methods=['POST'])
@admin_required
def toggle_status_familia(familia_id):
    """Ativa ou inativa uma família. Apenas admin."""
    familia = Familia.query.get_or_404(familia_id)
    if familia.status == 'inativa':
        familia.status = 'ativa'
        flash(f'Família de {familia.nome_responsavel} reativada com sucesso.', 'success')
    else:
        familia.status = 'inativa'
        flash(f'Família de {familia.nome_responsavel} inativada.', 'warning')
    db.session.commit()
    return redirect(request.referrer or url_for('listar_familias'))


@app.route('/admin/familias/cadastrar', methods=['GET', 'POST'])
@admin_required
def cadastrar_familia():
    """Cadastra uma nova família assistida. Apenas admin."""
    import json

    vicentinos = (
        Vicentino.query
        .filter_by(tipo='vicentino', status='ativo')
        .order_by(Vicentino.nome)
        .all()
    )

    vicentinos_data = [
        {
            'id': v.id,
            'nome': f"{v.nome} {v.sobrenome}",
            'conferencia_id': v.conferencia_id,
            'conferencia_nome': v.conferencia_rel.nome if v.conferencia_rel else ''
        }
        for v in vicentinos
    ]

    if request.method == 'POST':
        nome_responsavel = request.form.get('nome_responsavel', '').strip()
        cpf_input        = request.form.get('cpf_responsavel', '').strip()
        cpf              = re.sub(r'\D', '', cpf_input)
        telefone1        = request.form.get('telefone_principal', '').strip()
        telefone2        = request.form.get('telefone_secundario', '').strip()
        endereco         = request.form.get('endereco', '').strip()
        numero           = request.form.get('numero', '').strip()
        complemento      = request.form.get('complemento', '').strip()
        bairro           = request.form.get('bairro', '').strip()
        cidade           = request.form.get('cidade', '').strip()
        estado           = request.form.get('estado', '').strip().upper()
        cep              = request.form.get('cep', '').strip()
        qtd_moradores    = request.form.get('quantidade_moradores', '').strip()
        qtd_criancas     = request.form.get('quantidade_criancas', '').strip()
        vicentino_id     = request.form.get('vicentino_id', '').strip()
        conferencia_id   = request.form.get('conferencia_id', '').strip() or None
        observacoes      = request.form.get('observacoes', '').strip()

        form_data = dict(
            nome_responsavel=nome_responsavel, cpf_responsavel=cpf_input,
            telefone_principal=telefone1, telefone_secundario=telefone2,
            endereco=endereco, numero=numero, complemento=complemento,
            bairro=bairro, cidade=cidade, estado=estado, cep=cep,
            quantidade_moradores=qtd_moradores, quantidade_criancas=qtd_criancas,
            vicentino_id=vicentino_id, conferencia_id=conferencia_id,
            observacoes=observacoes,
        )

        def erro(msg):
            flash(msg, 'danger')
            return render_template(
                'admin_cadastrar_familia.html',
                vicentinos_data=json.dumps(vicentinos_data),
                vicentinos=vicentinos,
                form_data=form_data,
            )

        if not nome_responsavel or not endereco or not numero or not bairro or not cidade:
            return erro('Preencha todos os campos obrigatórios.')

        if not vicentino_id:
            return erro('Selecione um Vicentino Responsável.')

        if cpf:
            if not cpf_valido(cpf):
                return erro('CPF inválido.')
            if Familia.query.filter_by(cpf_responsavel=cpf).first():
                return erro('Já existe uma família cadastrada com esse CPF.')

        nova_familia = Familia(
            nome_responsavel  = nome_responsavel,
            cpf_responsavel   = cpf if cpf else None,
            telefone_principal  = telefone1 or None,
            telefone_secundario = telefone2 or None,
            endereco          = endereco,
            numero            = numero,
            complemento       = complemento or None,
            bairro            = bairro,
            cidade            = cidade,
            estado            = estado or None,
            cep               = cep or None,
            quantidade_moradores = int(qtd_moradores) if qtd_moradores.isdigit() else None,
            quantidade_criancas  = int(qtd_criancas)  if qtd_criancas.isdigit()  else None,
            vicentino_id      = int(vicentino_id),
            conferencia_id    = int(conferencia_id) if conferencia_id else None,
            observacoes       = observacoes or None,
            status            = 'ativa',
        )

        db.session.add(nova_familia)
        db.session.commit()

        flash('Família cadastrada com sucesso!', 'success')
        return redirect(url_for('dashboard'))

    return render_template(
        'admin_cadastrar_familia.html',
        vicentinos_data=json.dumps(vicentinos_data),
        vicentinos=vicentinos
    )


@app.route('/logout')
def logout():
    """Encerra a sessão do usuário e redireciona para a página de login."""
    session.clear()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))


# =========================================================
# CONFIRMAÇÃO DE E-MAIL
# =========================================================

@app.route('/confirmar_email/<token>')
def confirmar_email(token):
    """
    Valida o token de confirmação de e-mail recebido por link.
    Ativa a conta do usuário se o token for válido e ainda não tiver sido confirmado.
    """
    email = validar_token(token, 'confirmar-email', 86400)

    if not email:
        flash('Link de confirmação inválido ou expirado.', 'danger')
        return redirect(url_for('login'))

    vicentino = Vicentino.query.filter_by(email=email).first()

    if not vicentino:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('login'))

    if vicentino.email_confirmado:
        flash('Este e-mail já foi confirmado. Faça login.', 'info')
        return redirect(url_for('login'))

    # Ativa a conta e registra a data de confirmação
    vicentino.email_confirmado = True
    vicentino.status = 'ativo'
    vicentino.data_confirmacao = datetime.utcnow()
    db.session.commit()

    flash('E-mail confirmado com sucesso. Agora você já pode fazer login.', 'success')
    return redirect(url_for('login'))


@app.route('/reenviar_confirmacao', methods=['POST'])
def reenviar_confirmacao():
    """
    Reenvia o e-mail de confirmação de conta para o usuário.
    Funciona tanto para usuários logados quanto deslogados.
    Aplica controle de tempo (cooldown de 50 segundos) entre reenvios.
    """
    agora = int(time.time())
    tempo_espera = 50

    # Identifica o usuário logado ou pelo e-mail informado no formulário
    if 'vicentino_id' in session:
        usuario = Vicentino.query.get(session['vicentino_id'])
        if not usuario:
            flash('Usuário não encontrado. Faça login novamente.', 'danger')
            return redirect(url_for('login'))
    else:
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash('Informe seu e-mail.', 'danger')
            return redirect(url_for('login'))

        usuario = Vicentino.query.filter_by(email=email).first()
        if not usuario:
            # Mensagem genérica para evitar enumeração de e-mails
            flash('Se o e-mail estiver cadastrado, um novo link será enviado.', 'info')
            return redirect(url_for('login'))

    # Verifica se o e-mail já está confirmado
    if usuario.email_confirmado:
        session.pop('email_pendente_confirmacao', None)
        session.pop('ultimo_envio_confirmacao', None)
        flash('Seu e-mail já está confirmado.', 'info')
        return redirect(url_for('editar_perfil')) if 'vicentino_id' in session else redirect(url_for('login'))

    # Aplica cooldown entre reenvios
    ultimo_envio = session.get('ultimo_envio_confirmacao', 0)
    if agora - ultimo_envio < tempo_espera:
        restantes = tempo_espera - (agora - ultimo_envio)
        flash(f'Aguarde {restantes} segundos para reenviar o e-mail.', 'warning')
        return redirect(url_for('editar_perfil')) if 'vicentino_id' in session else redirect(url_for('login'))

    # Realiza o envio e atualiza o controle de tempo na sessão
    try:
        enviar_email_confirmacao(usuario)
        session['email_pendente_confirmacao'] = usuario.email
        session['ultimo_envio_confirmacao'] = agora
        flash('E-mail de confirmação reenviado com sucesso.', 'success')
    except Exception as e:
        flash(f'Erro ao reenviar e-mail: {str(e)}', 'danger')

    return redirect(url_for('editar_perfil')) if 'vicentino_id' in session else redirect(url_for('login'))


# =========================================================
# RECUPERAÇÃO DE SENHA
# =========================================================

@app.route('/esqueci_senha', methods=['GET', 'POST'])
def esqueci_senha():
    """
    Gerencia a recuperação e alteração de senha.
    - Usuário logado: pode alterar a senha diretamente ou solicitar link por e-mail.
    - Usuário deslogado: solicita link de redefinição via e-mail.
    Aplica cooldown de 50 segundos entre envios de e-mail.
    """
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        nova_senha = request.form.get('nova_senha', '').strip()
        confirmar_senha = request.form.get('confirmar_senha', '').strip()

        agora = int(time.time())
        tempo_espera = 50

        # Alteração direta de senha para usuário logado
        if 'vicentino_id' in session and nova_senha and confirmar_senha:
            vicentino = Vicentino.query.get(session['vicentino_id'])
            if not vicentino:
                flash('Usuário não encontrado. Faça login novamente.', 'danger')
                return redirect(url_for('login'))

            if nova_senha != confirmar_senha:
                flash('As senhas não coincidem.', 'danger')
                return redirect(url_for('esqueci_senha'))

            if len(nova_senha) < 6:
                flash('A senha deve ter pelo menos 6 caracteres.', 'danger')
                return redirect(url_for('esqueci_senha'))

            vicentino.senha_hash = generate_password_hash(nova_senha)
            db.session.commit()
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('dashboard'))

        # Solicita link por e-mail (usuário deslogado sem e-mail informado)
        if not email and 'vicentino_id' not in session:
            flash('Informe seu e-mail.', 'danger')
            return redirect(url_for('esqueci_senha'))

        # Busca o vicentino correto conforme o contexto (logado ou não)
        if 'vicentino_id' in session:
            vicentino = Vicentino.query.get(session['vicentino_id'])
            if not vicentino:
                flash('Usuário não encontrado. Faça login novamente.', 'danger')
                return redirect(url_for('login'))
        else:
            vicentino = Vicentino.query.filter_by(email=email).first()

        if vicentino:
            # Bloqueia redefinição se o e-mail ainda não foi confirmado
            if not vicentino.email_confirmado:
                flash('E-mail ainda não confirmado. Solicite ativação ou contate o administrador.', 'warning')
                return redirect(url_for('login'))

            # Aplica cooldown entre envios
            ultimo_envio = session.get('ultimo_envio_redefinicao', 0)
            if agora - ultimo_envio < tempo_espera:
                restantes = tempo_espera - (agora - ultimo_envio)
                flash(f'Aguarde {restantes} segundos para reenviar o e-mail de redefinição.', 'warning')
                return redirect(url_for('esqueci_senha'))

            # Envia e-mail de redefinição e registra na sessão
            try:
                enviar_email_redefinicao(vicentino)
                session['email_pendente_redefinicao'] = vicentino.email
                session['ultimo_envio_redefinicao'] = int(time.time())
                flash('Se o e-mail estiver cadastrado e confirmado, você receberá um link para redefinir a senha.',
                      'info')
            except Exception as e:
                flash(f'Erro ao enviar e-mail: {str(e)}', 'danger')
                return redirect(url_for('esqueci_senha'))
        else:
            # Resposta genérica para evitar enumeração de e-mails
            session['email_pendente_redefinicao'] = email
            session['ultimo_envio_redefinicao'] = int(time.time())
            flash('Se o e-mail estiver cadastrado, você receberá um link para redefinir a senha.', 'info')

        return redirect(url_for('esqueci_senha'))

    # GET: verifica tempo restante do cooldown e renderiza a página
    email_pendente = session.get('email_pendente_redefinicao')
    ultimo_envio = session.get('ultimo_envio_redefinicao', 0)
    agora = int(time.time())

    tempo_restante = max(0, 50 - (agora - ultimo_envio)) if email_pendente else 0

    # Limpa os dados de sessão se o cooldown já expirou
    if email_pendente and tempo_restante == 0:
        session.pop('email_pendente_redefinicao', None)
        session.pop('ultimo_envio_redefinicao', None)
        email_pendente = None

    return render_template(
        'esqueci_senha.html',
        email_pendente=email_pendente,
        tempo_restante=tempo_restante
    )


@app.route('/reenviar_redefinicao', methods=['POST'])
def reenviar_redefinicao():
    """
    Reenvia o e-mail de redefinição de senha.
    O e-mail pode vir do formulário ou da sessão ativa.
    Aplica cooldown de 50 segundos entre reenvios.
    """
    email = request.form.get('email', '').strip().lower()

    # Usa o e-mail da sessão como fallback
    if not email:
        email = session.get('email_pendente_redefinicao', '').strip().lower()

    if not email:
        flash('Informe seu e-mail.', 'danger')
        return redirect(url_for('esqueci_senha'))

    ultimo_envio = session.get('ultimo_envio_redefinicao', 0)
    agora = int(time.time())

    # Verifica o cooldown antes de reenviar
    if agora - ultimo_envio < 50:
        restantes = 50 - (agora - ultimo_envio)
        flash(f'Aguarde {restantes} segundos para reenviar o e-mail de redefinição.', 'warning')
        return redirect(url_for('esqueci_senha'))

    vicentino = Vicentino.query.filter_by(email=email).first()

    if vicentino:
        try:
            enviar_email_redefinicao(vicentino)
            session['email_pendente_redefinicao'] = email
            session['ultimo_envio_redefinicao'] = int(time.time())
            flash('E-mail de redefinição reenviado com sucesso.', 'success')
        except Exception as e:
            flash(f'Erro ao reenviar e-mail: {str(e)}', 'danger')
            return redirect(url_for('esqueci_senha'))
    else:
        # Resposta genérica para evitar enumeração de e-mails
        session['email_pendente_redefinicao'] = email
        session['ultimo_envio_redefinicao'] = int(time.time())
        flash('Se o e-mail estiver cadastrado, você receberá um link para redefinir a senha.', 'info')

    return redirect(url_for('esqueci_senha'))


@app.route('/redefinir_senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    """
    Exibe e processa o formulário de redefinição de senha via token.
    Valida o token, verifica os campos e atualiza a senha no banco.
    Limpa os dados de sessão relacionados à redefinição após o sucesso.
    """
    email = validar_token(token, 'redefinir-senha', 900)

    if not email:
        flash('Link inválido ou expirado.', 'danger')
        return redirect(url_for('login'))

    vicentino = Vicentino.query.filter_by(email=email).first()

    if not vicentino:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        senha = request.form.get('senha', '')
        confirmar_senha = request.form.get('confirmar_senha', '')

        if not senha or not confirmar_senha:
            flash('Preencha os dois campos de senha.', 'danger')
            return redirect(url_for('redefinir_senha', token=token))

        if len(senha) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'danger')
            return redirect(url_for('redefinir_senha', token=token))

        if senha != confirmar_senha:
            flash('As senhas não coincidem.', 'danger')
            return redirect(url_for('redefinir_senha', token=token))

        # Atualiza a senha e limpa os dados de sessão relacionados
        vicentino.senha_hash = generate_password_hash(senha)
        db.session.commit()

        session.pop('email_pendente_redefinicao', None)
        session.pop('ultimo_envio_redefinicao', None)
        session.pop('email_pendente_confirmacao', None)
        session.pop('ultimo_envio_confirmacao', None)

        flash('Senha redefinida com sucesso. Faça login.', 'success')
        return redirect(url_for('login'))

    return render_template('redefinir_senha.html', token=token)


# =========================================================
# ÁREA LOGADA
# =========================================================

@app.route('/dashboard')
def dashboard():
    """
    Exibe o painel principal do sistema.
    Redireciona para o login se o usuário não estiver autenticado.
    """
    if 'vicentino_id' not in session:
        flash('Faça login para acessar o sistema.', 'warning')
        return redirect(url_for('login'))

    return render_template('dashboard.html')

@app.route('/perfil')
def perfil():
    """Exibe a página de perfil do usuário logado."""
    if 'vicentino_id' not in session:
        return redirect(url_for('login'))

    usuario = Vicentino.query.get(session['vicentino_id'])
    if not usuario:
        session.clear()
        flash('Sessão inválida. Faça login novamente.', 'warning')
        return redirect(url_for('login'))

    return render_template('perfil.html', usuario=usuario)


@app.route('/editar_perfil', methods=['GET', 'POST'])
def editar_perfil():
    """
    Exibe e processa o formulário de edição de perfil do usuário logado.
    Permite atualizar nome, sobrenome, telefone, foto de perfil e senha.
    Administradores também podem editar conselho e conferência.
    Inclui reenvio de e-mail de confirmação com controle de cooldown.
    """
    if 'vicentino_id' not in session:
        flash('Faça login para acessar seu perfil.', 'warning')
        return redirect(url_for('login'))

    usuario = Vicentino.query.get(session['vicentino_id'])

    if not usuario:
        session.clear()
        flash('Usuário não encontrado. Faça login novamente.', 'danger')
        return redirect(url_for('login'))

    erros = {}
    conselhos = Conselho.query.all()
    extensoes_permitidas = {'jpg', 'jpeg', 'png', 'webp'}
    agora = int(time.time())
    ultimo_envio = session.get('ultimo_envio_confirmacao', 0)
    conselhos = Conselho.query.order_by(Conselho.nome).all()

    # Calcula o tempo restante do cooldown de reenvio de confirmação
    tempo_restante = max(0, 50 - (agora - ultimo_envio)) if usuario.email and not usuario.email_confirmado else 0

    if request.method == 'POST':

        # ── Reenvio de confirmação de e-mail ──────────────────────────────
        if 'reenviar_confirmacao' in request.form:
            if not usuario.email:
                flash('Você não possui e-mail cadastrado.', 'warning')
                return render_template(
                    'editar_perfil.html',
                    usuario=usuario,
                    erros=erros,
                    conselhos=conselhos,
                    telefone_input=telefone_para_input(usuario.telefone),
                    tempo_restante=0,
                )

            if usuario.email_confirmado:
                flash('Seu e-mail já está confirmado.', 'info')
                return render_template(
                    'editar_perfil.html',
                    usuario=usuario,
                    erros=erros,
                    conselhos=conselhos,
                    telefone_input=telefone_para_input(usuario.telefone),
                    tempo_restante=0,
                )

            if tempo_restante > 0:
                flash(f'Aguarde {tempo_restante} segundos para reenviar o e-mail.', 'warning')
                return render_template(
                    'editar_perfil.html',
                    usuario=usuario,
                    erros=erros,
                    conselhos=conselhos,
                    telefone_input=telefone_para_input(usuario.telefone),
                    tempo_restante=tempo_restante,
                )

            try:
                enviar_email_confirmacao(usuario)
                session['ultimo_envio_confirmacao'] = int(time.time())
                flash('E-mail de confirmação reenviado com sucesso.', 'success')
                tempo_restante = 50
            except Exception as e:
                flash(f'Erro ao reenviar e-mail: {str(e)}', 'danger')

            return render_template(
                'editar_perfil.html',
                usuario=usuario,
                erros=erros,
                conselhos=conselhos,
                telefone_input=telefone_para_input(usuario.telefone),
                tempo_restante=tempo_restante,
            )

        # ── Leitura dos campos do formulário ──────────────────────────────
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        telefone = request.form.get('telefone', '').strip()
        conselho_id = request.form.get('conselho_id') or None
        conferencia_id = request.form.get('conferencia_id') or None
        cpf_input = request.form.get('cpf', '').strip()
        documento = limpar_numero(cpf_input)
        senha_atual = request.form.get('senha_atual', '').strip()
        nova_senha = request.form.get('nova_senha', '').strip()
        confirmar_senha = request.form.get('confirmar_senha', '').strip()

        # ── Validação dos campos de texto ─────────────────────────────────
        if not nome:
            erros['nome'] = 'Informe o nome.'
        elif not apenas_letras(nome):
            erros['nome'] = 'O nome deve conter apenas letras.'

        if not sobrenome:
            erros['sobrenome'] = 'Informe o sobrenome.'
        elif not apenas_letras(sobrenome):
            erros['sobrenome'] = 'O sobrenome deve conter apenas letras.'

        # Valida o telefone removendo a máscara antes
        telefone_formatado = None
        if telefone:
            telefone_numeros = re.sub(r'\D', '', telefone)
            if telefone_numeros.startswith('55') and len(telefone_numeros) > 10:
                telefone_numeros = telefone_numeros[2:]
            telefone_formatado = telefone_valido_br(telefone_numeros)
            if not telefone_formatado:
                erros['telefone'] = 'Telefone inválido.'

        # ── Validação e processamento da foto de perfil ───────────────────
        foto = request.files.get('foto')
        imagem = None
        if foto and foto.filename:
            nome_foto = secure_filename(foto.filename)
            extensao = nome_foto.rsplit('.', 1)[-1].lower() if '.' in nome_foto else ''
            if extensao not in extensoes_permitidas:
                erros['foto'] = 'Formato de arquivo não suportado. Envie jpg, jpeg, png ou webp.'
            else:
                try:
                    foto.stream.seek(0)
                    imagem = Image.open(foto)
                    # Converte para RGB se necessário (ex: PNG com transparência)
                    if imagem.mode in ('RGBA', 'P'):
                        imagem = imagem.convert('RGB')
                    # Redimensiona para 400x400 px
                    imagem = imagem.resize((400, 400), Image.Resampling.LANCZOS)
                except UnidentifiedImageError:
                    erros['foto'] = 'Arquivo inválido. Envie apenas imagens.'
                except Exception:
                    erros['foto'] = 'Não foi possível processar a imagem.'

        # ── Validação e atualização de senha ──────────────────────────────
        if any([senha_atual, nova_senha, confirmar_senha]):
            # Troca de senha só é permitida para usuários com e-mail cadastrado
            if not usuario.email:
                flash('Você não possui email cadastrado. Procure a secretaria.', 'warning')
                return render_template(
                    'editar_perfil.html',
                    usuario=usuario,
                    erros=erros,
                    conselhos=conselhos,
                    telefone_input=telefone_para_input(usuario.telefone),
                    tempo_restante=tempo_restante,
                )

            if not senha_atual:
                erros['senha_atual'] = 'Informe a senha atual.'
            if not nova_senha:
                erros['nova_senha'] = 'Informe a nova senha.'
            if not confirmar_senha:
                erros['confirmar_senha'] = 'Confirme a nova senha.'
            if senha_atual and not check_password_hash(usuario.senha_hash, senha_atual):
                erros['senha_atual'] = 'Senha atual incorreta.'
            if nova_senha and confirmar_senha and nova_senha != confirmar_senha:
                erros['confirmar_senha'] = 'A confirmação da senha não coincide.'
            if nova_senha and len(nova_senha) < 6:
                erros['nova_senha'] = 'A nova senha deve ter pelo menos 6 caracteres.'
            # Aplica a nova senha apenas se todas as validações passaram
            if 'senha_atual' not in erros and 'nova_senha' not in erros and 'confirmar_senha' not in erros:
                usuario.senha_hash = generate_password_hash(nova_senha)

        # ── Retorna com erros se houver problemas de validação ────────────
        if erros:
            for campo, mensagem in erros.items():
                flash(mensagem, 'danger')
            return render_template(
                'editar_perfil.html',
                usuario=usuario,
                erros=erros,
                conselhos=conselhos,
                telefone_input=telefone_para_input(usuario.telefone),
                tempo_restante=tempo_restante,
            )

        # ── Atualiza os dados no banco ────────────────────────────────────
        usuario.nome = nome
        usuario.sobrenome = sobrenome
        usuario.telefone = telefone_formatado if telefone else None

        # Campos exclusivos para administradores
        if usuario.tipo == 'admin':
            usuario.conselho_id = int(conselho_id) if conselho_id else None
            usuario.conferencia_id = int(conferencia_id) if conferencia_id else None
            usuario.cnpj = documento if documento else None
            usuario.cpf = None
        else:
            usuario.cpf = documento if documento else None

        db.session.commit()

        # ── Salva a foto de perfil no disco ──────────────────────────────
        if imagem:
            try:
                nome_arquivo = f'vicentino_{usuario.id}.jpg'
                caminho_foto = os.path.join(app.config['UPLOAD_FOLDER_USUARIO'], nome_arquivo)

                # Remove a foto anterior se existir
                if usuario.foto:
                    caminho_antigo = os.path.join(app.config['UPLOAD_FOLDER_USUARIO'], usuario.foto)
                    if os.path.exists(caminho_antigo):
                        os.remove(caminho_antigo)

                imagem.save(caminho_foto, quality=90)
                usuario.foto = nome_arquivo
                db.session.commit()
            except Exception:
                flash('Perfil atualizado, mas houve erro ao salvar a imagem.', 'warning')

        # ── Atualiza os dados da sessão ───────────────────────────────────
        session['vicentino_nome'] = usuario.nome
        session['vicentino_email'] = usuario.email
        session['vicentino_foto'] = usuario.foto if usuario.foto else None

        flash('Perfil atualizado com sucesso!', 'success')
        return render_template(
            'editar_perfil.html',
            usuario=usuario,
            erros={},
            conselhos=conselhos,
            telefone_input=telefone_para_input(usuario.telefone),
            tempo_restante=tempo_restante,
        )

    # GET: renderiza a página com os dados atuais do usuário
    return render_template(
        'editar_perfil.html',
        usuario=usuario,
        erros=erros,
        conselhos=conselhos,
        telefone_input=telefone_para_input(usuario.telefone),
        tempo_restante=tempo_restante,
    )

@app.route('/admin/atendimentos')
@admin_required
def admin_atendimentos():
    """Lista todos os atendimentos registrados, com filtros. Apenas admin."""
    vicentino_raw = request.args.get('vicentino_id', '')
    familia_raw   = request.args.get('familia_id', '')
    data_inicio   = request.args.get('data_inicio', '')
    data_fim      = request.args.get('data_fim', '')

    query = Atendimento.query

    if vicentino_raw:
        try:
            query = query.filter_by(vicentino_id=int(vicentino_raw))
        except (ValueError, TypeError):
            pass

    if familia_raw:
        try:
            query = query.filter_by(familia_id=int(familia_raw))
        except (ValueError, TypeError):
            pass

    if data_inicio:
        try:
            query = query.filter(
                Atendimento.data_atendimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if data_fim:
        try:
            query = query.filter(
                Atendimento.data_atendimento <= datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    atendimentos = (
        query
        .order_by(Atendimento.data_atendimento.desc(), Atendimento.horario.desc())
        .all()
    )
    vicentinos   = Vicentino.query.filter_by(tipo='vicentino', status='ativo').order_by(Vicentino.nome).all()
    familias     = Familia.query.order_by(Familia.nome_responsavel).all()
    conferencias = Conferencia.query.order_by(Conferencia.nome).all()

    return render_template(
        'admin_atendimentos.html',
        atendimentos=atendimentos,
        vicentinos=vicentinos,
        familias=familias,
        conferencias=conferencias,
    )


@app.route('/admin/atendimentos/relatorio')
@admin_required
def relatorio_atendimentos():
    """Gera PDF de atendimentos com filtros independentes da listagem."""
    vicentino_raw    = request.args.get('rel_vicentino_id', '')
    conferencia_raw  = request.args.get('rel_conferencia_id', '')
    familia_raw      = request.args.get('rel_familia_id', '')
    data_inicio      = request.args.get('rel_data_inicio', '')
    data_fim         = request.args.get('rel_data_fim', '')

    query = Atendimento.query.join(Familia)

    if vicentino_raw:
        try:
            query = query.filter(Atendimento.vicentino_id == int(vicentino_raw))
        except (ValueError, TypeError):
            pass

    if conferencia_raw:
        try:
            query = query.filter(Familia.conferencia_id == int(conferencia_raw))
        except (ValueError, TypeError):
            pass

    if familia_raw:
        try:
            query = query.filter(Atendimento.familia_id == int(familia_raw))
        except (ValueError, TypeError):
            pass

    if data_inicio:
        try:
            query = query.filter(
                Atendimento.data_atendimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if data_fim:
        try:
            query = query.filter(
                Atendimento.data_atendimento <= datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    atendimentos = (
        query
        .order_by(Atendimento.data_atendimento.desc(), Atendimento.horario.desc())
        .all()
    )

    # Nomes para exibir no cabeçalho do PDF
    label_vicentino   = ''
    label_conferencia = ''
    label_familia     = ''
    if vicentino_raw:
        v = Vicentino.query.get(int(vicentino_raw))
        if v:
            label_vicentino = f"{v.nome} {v.sobrenome}"
    if conferencia_raw:
        c = Conferencia.query.get(int(conferencia_raw))
        if c:
            label_conferencia = c.nome
    if familia_raw:
        f = Familia.query.get(int(familia_raw))
        if f:
            label_familia = f.nome_responsavel

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    style_title   = ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold',
                                   alignment=TA_CENTER, spaceAfter=4)
    style_sub     = ParagraphStyle('sub', fontSize=9, fontName='Helvetica',
                                   alignment=TA_CENTER, textColor=colors.grey, spaceAfter=2)
    style_filtros = ParagraphStyle('filtros', fontSize=8, fontName='Helvetica',
                                   alignment=TA_CENTER, textColor=colors.HexColor('#444444'), spaceAfter=2)
    style_cell    = ParagraphStyle('cell', fontSize=8, fontName='Helvetica', leading=11)
    style_bold    = ParagraphStyle('bold', fontSize=8, fontName='Helvetica-Bold', leading=11)

    elements = []

    logo_path = os.path.join(app.static_folder, 'vicentino_2.jpg')
    if os.path.exists(logo_path):
        elements.append(RLImage(logo_path, width=3*cm, height=3*cm))
        elements.append(Spacer(1, 0.3*cm))

    elements.append(Paragraph('Relatório de Atendimentos', style_title))
    elements.append(Paragraph('Sociedade de São Vicente de Paulo', style_sub))
    elements.append(Spacer(1, 0.2*cm))

    # Filtros aplicados
    linhas_filtro = []
    if data_inicio or data_fim:
        periodo = ''
        if data_inicio:
            periodo += f"De {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        if data_fim:
            periodo += f"  até  {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        linhas_filtro.append(f"Período: {periodo.strip()}")
    if label_vicentino:
        linhas_filtro.append(f"Vicentino: {label_vicentino}")
    if label_conferencia:
        linhas_filtro.append(f"Conferência: {label_conferencia}")
    if label_familia:
        linhas_filtro.append(f"Família: {label_familia}")

    if linhas_filtro:
        elements.append(Paragraph(' · '.join(linhas_filtro), style_filtros))

    elements.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} · {len(atendimentos)} atendimento(s)",
        style_sub
    ))
    elements.append(Spacer(1, 0.5*cm))

    if not atendimentos:
        elements.append(Paragraph('Nenhum atendimento encontrado para os filtros selecionados.', styles['Normal']))
    else:
        header = [
            Paragraph('Família / Endereço', style_bold),
            Paragraph('Data / Horário', style_bold),
            Paragraph('Vicentino', style_bold),
            Paragraph('Descrição', style_bold),
            Paragraph('Itens doados', style_bold),
        ]
        rows = [header]

        for a in atendimentos:
            endereco = f"{a.familia.endereco}, {a.familia.numero}"
            if a.familia.complemento:
                endereco += f" – {a.familia.complemento}"
            endereco += f"\n{a.familia.bairro}, {a.familia.cidade}"

            data_hora = a.data_atendimento.strftime('%d/%m/%Y')
            if a.horario:
                data_hora += f"\n{a.horario.strftime('%H:%M')}"

            rows.append([
                Paragraph(f"<b>{a.familia.nome_responsavel}</b>\n{endereco}", style_cell),
                Paragraph(data_hora, style_cell),
                Paragraph(f"{a.vicentino.nome} {a.vicentino.sobrenome}", style_cell),
                Paragraph(a.descricao or '—', style_cell),
                Paragraph(a.itens_doados or '—', style_cell),
            ])

        col_widths = [4.5*cm, 2.5*cm, 3.5*cm, 4.5*cm, 3.5*cm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#0064B6')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F7FF')]),
            ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN',         (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING',    (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',   (0, 0), (-1, -1), 5),
            ('TOPPADDING',     (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_atendimentos_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    )
    return response


@app.route('/admin/atendimentos/relatorio/excel')
@admin_required
def relatorio_atendimentos_excel():
    """Gera Excel de atendimentos com os mesmos filtros do relatório PDF."""
    vicentino_raw   = request.args.get('rel_vicentino_id', '')
    conferencia_raw = request.args.get('rel_conferencia_id', '')
    familia_raw     = request.args.get('rel_familia_id', '')
    data_inicio     = request.args.get('rel_data_inicio', '')
    data_fim        = request.args.get('rel_data_fim', '')

    query = Atendimento.query.join(Familia)

    if vicentino_raw:
        try:
            query = query.filter(Atendimento.vicentino_id == int(vicentino_raw))
        except (ValueError, TypeError):
            pass
    if conferencia_raw:
        try:
            query = query.filter(Familia.conferencia_id == int(conferencia_raw))
        except (ValueError, TypeError):
            pass
    if familia_raw:
        try:
            query = query.filter(Atendimento.familia_id == int(familia_raw))
        except (ValueError, TypeError):
            pass
    if data_inicio:
        try:
            query = query.filter(
                Atendimento.data_atendimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass
    if data_fim:
        try:
            query = query.filter(
                Atendimento.data_atendimento <= datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    atendimentos = (
        query
        .order_by(Atendimento.data_atendimento.desc(), Atendimento.horario.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Atendimentos'

    azul        = '0064B6'
    azul_claro  = 'E8F1FB'
    branco      = 'FFFFFF'
    cinza_borda = 'CCCCCC'

    borda = Border(
        left=Side(style='thin', color=cinza_borda),
        right=Side(style='thin', color=cinza_borda),
        top=Side(style='thin', color=cinza_borda),
        bottom=Side(style='thin', color=cinza_borda),
    )

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório de Atendimentos — SSVP'
    ws['A1'].font      = Font(bold=True, size=14, color=azul)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} · {len(atendimentos)} atendimento(s)"
    ws['A2'].font      = Font(size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # Cabeçalho da tabela
    cabecalhos = ['Família', 'Endereço', 'Bairro / Cidade', 'Conferência', 'Data', 'Horário', 'Vicentino', 'Descrição', 'Itens Doados']
    ws.append([])  # linha 3 vazia
    ws.append(cabecalhos)  # linha 4

    for col_idx, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font      = Font(bold=True, color=branco, size=10)
        cell.fill      = PatternFill(fill_type='solid', fgColor=azul)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = borda
    ws.row_dimensions[4].height = 20

    # Dados
    for i, a in enumerate(atendimentos):
        endereco = f"{a.familia.endereco}, {a.familia.numero}"
        if a.familia.complemento:
            endereco += f" – {a.familia.complemento}"

        linha = [
            a.familia.nome_responsavel,
            endereco,
            f"{a.familia.bairro} – {a.familia.cidade}",
            a.familia.conferencia_rel.nome if a.familia.conferencia_rel else '',
            a.data_atendimento.strftime('%d/%m/%Y'),
            a.horario.strftime('%H:%M') if a.horario else '',
            f"{a.vicentino.nome} {a.vicentino.sobrenome}",
            a.descricao or '',
            a.itens_doados or '',
        ]
        row_num = 5 + i
        ws.append(linha)
        fill_color = branco if i % 2 == 0 else azul_claro
        for col_idx in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill      = PatternFill(fill_type='solid', fgColor=fill_color)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border    = borda
        ws.row_dimensions[row_num].height = 40

    # Largura das colunas
    larguras = [28, 30, 25, 22, 12, 10, 25, 40, 35]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_atendimentos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    return response


@app.route('/admin/atendimentos/relatorio/csv')
@admin_required
def relatorio_atendimentos_csv():
    """Gera CSV de atendimentos com os mesmos filtros do relatório."""
    vicentino_raw   = request.args.get('rel_vicentino_id', '')
    conferencia_raw = request.args.get('rel_conferencia_id', '')
    familia_raw     = request.args.get('rel_familia_id', '')
    data_inicio     = request.args.get('rel_data_inicio', '')
    data_fim        = request.args.get('rel_data_fim', '')

    query = Atendimento.query.join(Familia)

    if vicentino_raw:
        try:
            query = query.filter(Atendimento.vicentino_id == int(vicentino_raw))
        except (ValueError, TypeError):
            pass
    if conferencia_raw:
        try:
            query = query.filter(Familia.conferencia_id == int(conferencia_raw))
        except (ValueError, TypeError):
            pass
    if familia_raw:
        try:
            query = query.filter(Atendimento.familia_id == int(familia_raw))
        except (ValueError, TypeError):
            pass
    if data_inicio:
        try:
            query = query.filter(
                Atendimento.data_atendimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass
    if data_fim:
        try:
            query = query.filter(
                Atendimento.data_atendimento <= datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    atendimentos = (
        query
        .order_by(Atendimento.data_atendimento.desc(), Atendimento.horario.desc())
        .all()
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')

    writer.writerow(['Família', 'Endereço', 'Bairro / Cidade', 'Conferência',
                     'Data', 'Horário', 'Vicentino', 'Descrição', 'Itens Doados'])

    for a in atendimentos:
        endereco = f"{a.familia.endereco}, {a.familia.numero}"
        if a.familia.complemento:
            endereco += f" – {a.familia.complemento}"
        writer.writerow([
            a.familia.nome_responsavel,
            endereco,
            f"{a.familia.bairro} – {a.familia.cidade}",
            a.familia.conferencia_rel.nome if a.familia.conferencia_rel else '',
            a.data_atendimento.strftime('%d/%m/%Y'),
            a.horario.strftime('%H:%M') if a.horario else '',
            f"{a.vicentino.nome} {a.vicentino.sobrenome}",
            a.descricao or '',
            a.itens_doados or '',
        ])

    response = make_response('﻿' + buffer.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = (
        f'attachment; filename="relatorio_atendimentos_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    )
    return response


# =============================
@app.route('/admin/vicentinos')
@admin_required
def listar_vicentinos():
    conselho_id_raw = request.args.get('conselho_id')
    conferencia_id_raw = request.args.get('conferencia_id')
    nome_busca = request.args.get('nome', '').strip()
    status_filtro = request.args.get('status', '')

    query = Vicentino.query.filter(Vicentino.tipo != 'admin')

    if nome_busca:
        termo = f'%{nome_busca}%'
        query = query.filter(
            db.or_(
                Vicentino.nome.ilike(termo),
                Vicentino.sobrenome.ilike(termo)
            )
        )

    if conselho_id_raw:
        try:
            query = query.filter_by(conselho_id=int(conselho_id_raw))
        except (ValueError, TypeError):
            flash('Filtro de conselho inválido.', 'danger')
            return redirect(url_for('listar_vicentinos'))

    if conferencia_id_raw:
        try:
            query = query.filter_by(conferencia_id=int(conferencia_id_raw))
        except (ValueError, TypeError):
            flash('Filtro de conferência inválido.', 'danger')
            return redirect(url_for('listar_vicentinos'))

    todos = query.all()

    if status_filtro == 'ativo':
        ativos = [v for v in todos if v.status != 'inativo']
        inativos = []
    elif status_filtro == 'inativo':
        ativos = []
        inativos = [v for v in todos if v.status == 'inativo']
    else:
        ativos = [v for v in todos if v.status != 'inativo']
        inativos = [v for v in todos if v.status == 'inativo']

    conselhos = Conselho.query.all()
    conferencias = Conferencia.query.all()

    return render_template(
        'admin_vicentinos.html',
        ativos=ativos,
        inativos=inativos,
        conselhos=conselhos,
        conferencias=conferencias
    )

@app.route('/admin/vicentino/<int:vicentino_id>/toggle_status', methods=['POST'])
@admin_required
def toggle_status_vicentino(vicentino_id):
    vicentino = Vicentino.query.get_or_404(vicentino_id)
    if vicentino.status == 'inativo':
        vicentino.status = 'ativo'
        flash(f'{vicentino.nome} foi reativado com sucesso.', 'success')
    else:
        vicentino.status = 'inativo'
        flash(f'{vicentino.nome} foi inativado.', 'warning')
    db.session.commit()
    return redirect(request.referrer or url_for('listar_vicentinos'))


@app.route('/admin/editar_vicentino/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_vicentino_admin(id):
    usuario = Vicentino.query.get_or_404(id)

    conselhos = Conselho.query.all()
    conferencias = Conferencia.query.all()

    erros = {}

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        telefone = request.form.get('telefone', '').strip()

        conselho_id = request.form.get('conselho_id') or None
        conferencia_id = request.form.get('conferencia_id') or None

        # =============================
        # VALIDAÇÕES
        # =============================

        if not nome or not apenas_letras(nome):
            erros['nome'] = 'Nome inválido.'

        if not sobrenome or not apenas_letras(sobrenome):
            erros['sobrenome'] = 'Sobrenome inválido.'

        # Telefone
        telefone_formatado = None
        if telefone:
            telefone_numeros = re.sub(r'\D', '', telefone)

            if telefone_numeros.startswith('55') and len(telefone_numeros) > 10:
                telefone_numeros = telefone_numeros[2:]

            telefone_formatado = telefone_valido_br(telefone_numeros)

            if not telefone_formatado:
                erros['telefone'] = 'Telefone inválido.'

        # =============================
        # SE TIVER ERRO
        # =============================
        if erros:
            for msg in erros.values():
                flash(msg, 'danger')

            return render_template(
                'admin_editar_vicentino.html',
                usuario=usuario,
                conselhos=conselhos,
                conferencias=conferencias
            )

        # =============================
        # ATUALIZAÇÃO
        # =============================
        usuario.nome = nome
        usuario.sobrenome = sobrenome
        usuario.telefone = telefone_formatado if telefone else None
        usuario.conselho_id = conselho_id
        usuario.conferencia_id = conferencia_id

        db.session.commit()

        flash('Vicentino atualizado com sucesso!', 'success')
        return redirect(url_for('listar_vicentinos'))

    return render_template(
        'admin_editar_vicentino.html',
        usuario=usuario,
        conselhos=conselhos,
        conferencias=conferencias
    )

@app.route('/admin/familia/<int:familia_id>/editar', methods=['GET', 'POST'])
@admin_required
def editar_familia(familia_id):
    """Edita os dados de uma família assistida. Apenas admin."""
    import json

    familia = Familia.query.get_or_404(familia_id)

    vicentinos = (
        Vicentino.query
        .filter_by(tipo='vicentino', status='ativo')
        .order_by(Vicentino.nome)
        .all()
    )

    vicentinos_data = [
        {
            'id': v.id,
            'nome': f"{v.nome} {v.sobrenome}",
            'conferencia_id': v.conferencia_id,
            'conferencia_nome': v.conferencia_rel.nome if v.conferencia_rel else ''
        }
        for v in vicentinos
    ]

    if request.method == 'POST':
        nome_responsavel = request.form.get('nome_responsavel', '').strip()
        cpf_input        = request.form.get('cpf_responsavel', '').strip()
        cpf              = re.sub(r'\D', '', cpf_input)
        telefone1        = request.form.get('telefone_principal', '').strip()
        telefone2        = request.form.get('telefone_secundario', '').strip()
        endereco         = request.form.get('endereco', '').strip()
        numero           = request.form.get('numero', '').strip()
        complemento      = request.form.get('complemento', '').strip()
        bairro           = request.form.get('bairro', '').strip()
        cidade           = request.form.get('cidade', '').strip()
        estado           = request.form.get('estado', '').strip().upper()
        cep              = request.form.get('cep', '').strip()
        qtd_moradores    = request.form.get('quantidade_moradores', '').strip()
        qtd_criancas     = request.form.get('quantidade_criancas', '').strip()
        vicentino_id     = request.form.get('vicentino_id', '').strip()
        conferencia_id   = request.form.get('conferencia_id', '').strip() or None
        observacoes      = request.form.get('observacoes', '').strip()
        status           = request.form.get('status', 'ativa')

        def erro(msg):
            flash(msg, 'danger')
            return render_template(
                'admin_editar_familia.html',
                familia=familia,
                vicentinos_data=json.dumps(vicentinos_data),
                vicentinos=vicentinos,
            )

        if not nome_responsavel or not endereco or not numero or not bairro or not cidade:
            return erro('Preencha todos os campos obrigatórios.')

        if not vicentino_id:
            return erro('Selecione um Vicentino Responsável.')

        if cpf:
            if not cpf_valido(cpf):
                return erro('CPF inválido.')
            duplicado = Familia.query.filter(
                Familia.cpf_responsavel == cpf,
                Familia.id != familia_id
            ).first()
            if duplicado:
                return erro('Já existe outra família cadastrada com esse CPF.')

        familia.nome_responsavel    = nome_responsavel
        familia.cpf_responsavel     = cpf if cpf else None
        familia.telefone_principal  = telefone1 or None
        familia.telefone_secundario = telefone2 or None
        familia.endereco            = endereco
        familia.numero              = numero
        familia.complemento         = complemento or None
        familia.bairro              = bairro
        familia.cidade              = cidade
        familia.estado              = estado or None
        familia.cep                 = cep or None
        familia.quantidade_moradores = int(qtd_moradores) if qtd_moradores.isdigit() else None
        familia.quantidade_criancas  = int(qtd_criancas)  if qtd_criancas.isdigit()  else None
        familia.vicentino_id        = int(vicentino_id)
        familia.conferencia_id      = int(conferencia_id) if conferencia_id else None
        familia.observacoes         = observacoes or None
        familia.status              = status

        db.session.commit()

        flash('Família atualizada com sucesso!', 'success')
        return redirect(url_for('listar_familias'))

    return render_template(
        'admin_editar_familia.html',
        familia=familia,
        vicentinos_data=json.dumps(vicentinos_data),
        vicentinos=vicentinos,
    )


@app.route('/minhas_familias')
def minhas_familias():
    """Lista as famílias atribuídas ao vicentino logado, com filtros."""
    if 'vicentino_id' not in session:
        flash('Faça login para acessar o sistema.', 'warning')
        return redirect(url_for('login'))
    if session.get('tipo') == 'admin':
        return redirect(url_for('dashboard'))

    vicentino_id  = session['vicentino_id']
    nome_busca    = request.args.get('nome', '').strip()
    bairro_busca  = request.args.get('bairro', '').strip()
    status_filtro = request.args.get('status', '')

    query = Familia.query.filter_by(vicentino_id=vicentino_id)

    if nome_busca:
        query = query.filter(Familia.nome_responsavel.ilike(f'%{nome_busca}%'))
    if bairro_busca:
        query = query.filter(Familia.bairro.ilike(f'%{bairro_busca}%'))

    todas = query.order_by(Familia.nome_responsavel).all()

    if status_filtro == 'ativa':
        ativas   = [f for f in todas if f.status == 'ativa']
        inativas = []
    elif status_filtro == 'inativa':
        ativas   = []
        inativas = [f for f in todas if f.status == 'inativa']
    else:
        ativas   = [f for f in todas if f.status == 'ativa']
        inativas = [f for f in todas if f.status == 'inativa']

    return render_template('minhas_familias.html', ativas=ativas, inativas=inativas)


@app.route('/meus_atendimentos')
def meus_atendimentos():
    """Lista os atendimentos registrados pelo vicentino logado, com filtros."""
    if 'vicentino_id' not in session:
        flash('Faça login para acessar o sistema.', 'warning')
        return redirect(url_for('login'))
    if session.get('tipo') == 'admin':
        return redirect(url_for('dashboard'))

    vicentino_id = session['vicentino_id']
    familia_raw  = request.args.get('familia_id', '')
    data_inicio  = request.args.get('data_inicio', '')
    data_fim     = request.args.get('data_fim', '')

    query = Atendimento.query.filter_by(vicentino_id=vicentino_id)

    if familia_raw:
        try:
            query = query.filter_by(familia_id=int(familia_raw))
        except (ValueError, TypeError):
            pass

    if data_inicio:
        try:
            query = query.filter(
                Atendimento.data_atendimento >= datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    if data_fim:
        try:
            query = query.filter(
                Atendimento.data_atendimento <= datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    atendimentos = (
        query
        .order_by(Atendimento.data_atendimento.desc(), Atendimento.horario.desc())
        .all()
    )
    familias = (
        Familia.query
        .filter_by(vicentino_id=vicentino_id)
        .order_by(Familia.nome_responsavel)
        .all()
    )

    return render_template('meus_atendimentos.html', atendimentos=atendimentos, familias=familias)


@app.route('/registrar_atendimento', methods=['GET', 'POST'])
def registrar_atendimento():
    """Formulário para o vicentino registrar um novo atendimento a uma família."""
    if 'vicentino_id' not in session:
        flash('Faça login para acessar o sistema.', 'warning')
        return redirect(url_for('login'))
    if session.get('tipo') == 'admin':
        return redirect(url_for('dashboard'))

    vicentino_id = session['vicentino_id']
    familias = Familia.query.filter_by(vicentino_id=vicentino_id, status='ativa').all()

    if request.method == 'POST':
        familia_id  = request.form.get('familia_id', '').strip()
        data_str    = request.form.get('data_atendimento', '').strip()
        horario_str = request.form.get('horario', '').strip()
        descricao   = request.form.get('descricao', '').strip()
        itens_doados = request.form.get('itens_doados', '').strip()

        if not familia_id or not data_str:
            flash('Família e data são obrigatórios.', 'danger')
            return render_template('registrar_atendimento.html', familias=familias, hoje=datetime.today().strftime('%Y-%m-%d'))

        familia = Familia.query.filter_by(id=familia_id, vicentino_id=vicentino_id).first()
        if not familia:
            flash('Família inválida.', 'danger')
            return render_template('registrar_atendimento.html', familias=familias, hoje=datetime.today().strftime('%Y-%m-%d'))

        try:
            data_atendimento = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data inválida.', 'danger')
            return render_template('registrar_atendimento.html', familias=familias, hoje=datetime.today().strftime('%Y-%m-%d'))

        horario = None
        if horario_str:
            try:
                horario = datetime.strptime(horario_str, '%H:%M').time()
            except ValueError:
                pass

        atendimento = Atendimento(
            familia_id=int(familia_id),
            vicentino_id=vicentino_id,
            data_atendimento=data_atendimento,
            horario=horario,
            descricao=descricao or None,
            itens_doados=itens_doados or None,
        )
        db.session.add(atendimento)
        db.session.commit()

        flash('Atendimento registrado com sucesso!', 'success')
        return redirect(url_for('meus_atendimentos'))

    return render_template('registrar_atendimento.html', familias=familias, hoje=datetime.today().strftime('%Y-%m-%d'))


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)